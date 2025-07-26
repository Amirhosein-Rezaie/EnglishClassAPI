from . import models
from . import serializers
from rest_framework.viewsets import ModelViewSet
from EnglishClass.permissions import (DeleteForAdmin, NotAllow)
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from EnglishClass.helper import dynamic_search, description_search_swagger
from drf_spectacular.utils import extend_schema
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from django.http import HttpResponse
from io import BytesIO
from education.helper import return_value_cell_term_excel
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from people.models import Students
from people.serializers import StudentSerializer
from EnglishClass.permissions import IsAdminOrReadOnly
from rest_framework.decorators import permission_classes


# terms
class TermViewset(ModelViewSet):
    serializer_class = serializers.TermSerializer
    queryset = models.Terms.objects.all()
    permission_classes = [DeleteForAdmin]

    @extend_schema(
        description=description_search_swagger
    )
    def list(self, request: Request, *args, **kwargs):
        if request.query_params:
            return dynamic_search(
                request=request, model=models.Terms,
                serializer=serializers.TermSerializer
            )
        return super().list(request, *args, **kwargs)


# count the number students in term with students data
@permission_classes([IsAdminOrReadOnly])
class term_students(APIView):
    def get(self, request: Request):
        query_params = request.query_params
        if not query_params.get('title') and not query_params.get('id'):
            return Response({
                "error": "no query_params to search",
                "detail": "پارامتری برای جست و جو پیدا نشد ... !",
            }, status=status.HTTP_400_BAD_REQUEST)
        id = query_params.get('id')

        result = {}
        result['count'] = Students.objects.filter(Q(
            id__in=models.Registers.objects.filter(
                Q(term=id)).values_list('student', flat=True)
        )).count()

        result['students'] = StudentSerializer(Students.objects.filter(
            id__in=models.Registers.objects.filter(
                term=id).values_list('student', flat=True)
        ).distinct(), many=True).data

        return Response(result, status=status.HTTP_200_OK)


# register
class RegisterViewset(ModelViewSet):
    serializer_class = serializers.RegisterSerializer
    queryset = models.Registers.objects.all()
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.request.method in ['DELETE']:
            return [NotAllow()]
        return super().get_permissions()

    @extend_schema(
        description=description_search_swagger
    )
    def list(self, request: Request, *args, **kwargs):
        if request.query_params:
            return dynamic_search(
                request=request, model=models.Registers,
                serializer=serializers.RegisterSerializer
            )
        return super().list(request, *args, **kwargs)


# grades
class GradeViewset(ModelViewSet):
    serializer_class = serializers.GradeSerializer
    queryset = models.Grades.objects.all()
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.request.method in ['DELETE']:
            return [NotAllow()]
        return super().get_permissions()

    @extend_schema(
        description=description_search_swagger
    )
    def list(self, request: Request, *args, **kwargs):
        if request.query_params:
            return dynamic_search(
                request=request, model=models.Grades,
                serializer=serializers.GradeSerializer
            )
        return super().list(request, *args, **kwargs)


# book sales
class BookSaleViewset(ModelViewSet):
    serializer_class = serializers.BookSaleSerializer
    queryset = models.BookSales.objects.all()
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.request.method in ['DELETE']:
            return [NotAllow()]
        return super().get_permissions()

    @extend_schema(
        description=description_search_swagger
    )
    def list(self, request: Request, *args, **kwargs):
        if request.query_params:
            return dynamic_search(
                request=request, model=models.BookSales,
                serializer=serializers.BookSaleSerializer
            )
        return super().list(request, *args, **kwargs)


# export excel file of terms
class export_terms_excel(APIView):
    def get(self, request: Request):
        wb = Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        ws.title = 'terms'

        # general style
        # # border
        side = Side(style='thin')
        border = Border(right=side, bottom=side, top=side)
        # # aligment
        alignment = Alignment(horizontal='center', vertical='center')
        # # font
        font = Font(size=14, name='Calibri')
        # # fill
        fill = PatternFill(fill_type='solid', fgColor='fdf2e7')
        # # cell height and width
        height = 25

        # header style
        # # fill
        header_fill = PatternFill(fill_type='solid', fgColor='daf6ff')
        # # cell height and width
        height = 30
        width = 40

        # header
        columns = [
            'ردیف', 'عنوان', 'کتاب کار', 'کتاب داستان',
            'کتاب زبان اموز', 'نام معلم',
            'شهریه', 'تاریخ شروع', 'تاریخ پایان',
            'ساعت شروع', 'ساعت پایان', 'نوع', 'سطح'
        ]
        for col in range(len(columns)):
            cell = ws.cell(row=1, column=col+1)
            cell.value = columns[col]
            cell.fill = header_fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width
            ws.row_dimensions[1].height = height

        # data
        terms = models.Terms.objects.all()
        row = 2
        for term in terms:
            for col in range(len(columns)):
                cell = ws.cell(row=row, column=col+1)
                cell.value = return_value_cell_term_excel(
                    col=col, queryset=term, row=row)
                cell.fill = fill if col != 0 else header_fill
                cell.alignment = alignment
                cell.border = border
            row += 1

        # save the file
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        # send the file
        response = HttpResponse(
            stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="terms.xlsx"'
        return response


# export excel file for grades
class export_grades_excel(APIView):
    def get(self, request: Request):
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = 'نمره ها'
        work_sheet.sheet_view.rightToLeft = True

        # genral style
        # # fill
        fill = PatternFill(fill_type='solid', fgColor='fffdee')
        # # border
        side = Side(style='thin')
        border = Border(right=side, top=side, bottom=side)
        # # font
        font = Font(name='Calibri', size=14, bold=True)
        # # sizes
        width = 20
        heigth = 25
        # # alignment
        alignment = Alignment(horizontal='center', vertical='center')

        # header style
        header_fill = PatternFill(fill_type='solid', fgColor='ccebff')
        header_font = Font(name='Calibri', size=14, bold=True)

        # data style
        total_grade_font_fail = Font(color='a91811', size=14, name='Calibri')
        total_grade_font_pass = Font(color='17964c', size=14, name='Calibri')
        grade_font_normal = Font(color='000000', name='calibri', size=14)

        # header
        columns = [
            'ردیف', 'زبان آموز', 'ترم',
            'نمره کلاسی', 'نمره کتاب کار', 'نمره کتاب داستان',
            'نمره کلیپ', 'نمره فیلم', 'نمره امتحان', 'نمره کل'
        ]
        for col in range(len(columns)):
            cell = work_sheet.cell(row=1, column=col+1)
            cell.value = columns[col]
            work_sheet.column_dimensions[cell.column_letter].width = width
            work_sheet.row_dimensions[1].height = heigth
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment
            cell.font = header_font

        # data
        graders = models.Grades.objects.all()
        row = 2
        for grader in graders:
            values = [
                row - 1,
                f"{grader.student.first_name} {grader.student.last_name}", f"{grader.term.title}",
                grader.class_grade, grader.workbook_grade, grader.Storybook_grade,
                grader.Videoclip_grade, grader.Film_grade,
                grader.Exam_grade, sum([
                    grader.class_grade, grader.workbook_grade, grader.Storybook_grade,
                    grader.Videoclip_grade, grader.Film_grade,
                    grader.Exam_grade
                ])
            ]
            for col in range(len(columns)):
                cell = work_sheet.cell(row=row, column=col+1)
                cell.value = values[col]
                cell.fill = fill if col != 0 else header_fill
                cell.border = border
                cell.font = grade_font_normal
                if col == len(columns) - 1:
                    if cell.value >= 65:
                        cell.font = total_grade_font_pass
                    else:
                        cell.font = total_grade_font_fail
                cell.alignment = alignment
            row += 1

        # save the file
        stream = BytesIO()
        work_book.save(stream)
        stream.seek(0)

        # send the file
        response = HttpResponse(
            stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="grades.xlsx"'
        return response
