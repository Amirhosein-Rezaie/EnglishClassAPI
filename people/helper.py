from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from django.db.models import Model
from rest_framework.serializers import ModelSerializer


def export_excel(model: Model, serializer: ModelSerializer, columns: list, filename: str):
    """
    a funstion that you can make a simple excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = filename
    ws.sheet_view.rightToLeft = True

    # general style
    # # alignment
    custom_alignment = Alignment(horizontal='center', vertical='center')
    # # border
    side = Side(style='thin')
    custom_border = Border(left=side, right=side, top=side, bottom=side,)

    # style headers or columns
    header_font = Font(name='Calibri', size=14)
    header_fill = PatternFill(fill_type='solid', fgColor='7ecbec')
    columns = ['ردیف'] + columns

    # style for datas
    data_fill = PatternFill(fill_type='solid', fgColor='f8eed9')
    data_row_fill = PatternFill(fill_type='solid', fgColor='f3bfab')

    # make header or columns
    for col in range(len(columns)):
        cell = ws.cell(row=1, column=col + 1)
        cell.value = columns[col]
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 20
        cell.alignment = custom_alignment
        cell.border = custom_border

    # make rows
    datas = serializer(
        model.objects.all(), many=True).data
    row = 2
    for data in datas:
        values = [value for value in dict(data).values()
                  if not isinstance(value, list)]
        for col in range(len(columns)):
            cell = ws.cell(row=row, column=col + 1)
            cell.value = values[col] if col != 0 else row - 1
            cell.border = custom_border
            cell.alignment = custom_alignment
            cell.fill = data_row_fill if col == 0 else data_fill
        row += 1

    # save the file
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # send the file
    response = HttpResponse(
        stream.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response
